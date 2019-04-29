#!/usr/bin/env python3
# -*- coding:utf-8 -*- 

'''
 @Author      : Simon Chen
 @Email       : bafelem@gmail.com
 @datetime    : 2019-03-05 16:47:31
 @Description : Description
 @FileName    : PFprocess.py
'''

import datetime
import os
import re
import shutil
import sys
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side

CRED = '\033[91m'
CEND = '\033[0m'
CGRN = '\033[0;32m'

HOME = os.path.expanduser('~')
template_path = HOME + '/Documents/DFtemp/' + 'template.xlsx'
repeat_template = HOME + '/Documents/DFtemp/' + 'template1.xlsx'

def run():
    if not os.path.exists(repeat_template):
        shutil.copyfile(template_path, repeat_template)
    global job
    while True:
        job = input("Please input job number: ")
        job = job.upper()
        if len(job) > 10:
            break
        else:
            print("Job number should not be less than 11 digits")
    fill_sheet()
    create_folder()
    print(CGRN + "Completed! Enjoy!" + CEND)
    sys.exit()

def create_folder():
    folder = job + ' 做稿'
    sub_folder1 = "1 intake sheet & order"
    sub_folder2 = "2 raw client files"
    time = datetime.datetime.now()
    today = datetime.date.today()
    time = int(time.hour)
    if time >= 7:
        date_folder = today.strftime( '%m%d')
    else:
        delta_time = today + datetime.timedelta( days = -1 )
        date_folder = delta_time.strftime( '%m%d')
    sub_folder1 = HOME + '/Desktop/' + folder + '/' + sub_folder1 + '/' + date_folder
    sub_folder2 = HOME + '/Desktop/' + folder + '/' + sub_folder2 + '/' + date_folder
    os.makedirs(sub_folder1)
    os.makedirs(sub_folder2)
    final_name = HOME + '/Desktop/' + folder +'/' + job + '_DetailList_W.xlsx'
    shutil.move(repeat_template, final_name)
    return

def fill_sheet():
    path1 = repeat_template
    h2 = job
    h2 = h2.upper()
    b31 = input("Please enter the season code:")
    b31 = b31.upper()
    b32 = input("Please enter country, A. China, B. USA, C.Canada, otherwise China:")
    if re.match('A', b32, re.IGNORECASE ):
        b32 = 'China'
    elif re.match('B', b32, re.IGNORECASE ):
        b32 = 'USA'
    elif re.match('C', b32, re.IGNORECASE ):
        b32 = 'Canada'
    else:
        b32 = 'China'
        print("It will default to China.")

    B3 = b31 + '/' + b32

    A5 = input("Please enter the date or default today:")
    if A5.strip() == "":
        time = datetime.datetime.now()
        time = int(time.hour)
        today = datetime.date.today()
        if time >= 7:
            A5 = today.strftime( '%m/%d/%Y')
        else:
            A5 = today + datetime.timedelta( days = -1 )
            A5 = A5.strftime( '%m/%d/%Y')

    dic = {'THX':'Thanksgiving', 'AW':'Athletic Works', 'NY':'New Year', 'HOL':'Holiday Time', 'EA':'Easter', 'SUM':'Summer', \
'SP':'Spring', 'CASE':'Casemate', 'VAL':'Valentines', 'ONN':'ONN', 'HV':'Harvest', 'HW':'Halloween', 'PD':'Play Day', 'GD':'Graduation', \
'MO':'Mother\'s Day', 'AMR':'Americana', 'GV':'Great Value', 'STP':'St. Patrick', 'MG':'Mardi Gras', 'BW':'Blackweb', 'HAN':'Hanukkah', \
'WTC':'Way to Celebrate', 'FA':'Father\'s Day', 'PC':'Parent\'s Choice', 'CA':'Canadiana', 'EQ':'Equate', 'MOV':'Movelo', \
'WTCW':'Way To Celebrate Wedding', 'OZ':'Ozark Trail', 'PA':'Patriotic'}

    try:
        code = b31
        full_name = dic[code]
    except KeyError:
        full_name = input("Please enter full season name:")
    finally:
        H5 = full_name

    H6 = input("Please enter supplier name:")
    H6 = H6.title()

    H7 = input("Please enter buyer name:")
    H7 = H7.title()
    H71 = input("Please enter Department number:")

    H11 = input("In-store date:")
    H10 = input("Shipdate:")
    H9 = input("Packout date:")
    H8 = input("Please Artwork due date:")
    H13 = input("Contact info:")

    wb = load_workbook(path1)
    ws = wb.active
    '''
    cell borders
    '''
    top_right_thin_border = Border(top=Side(style='thin'),
                                  right=Side(style='thin'))
    top_thin_border = Border(top=Side(style='thin'))
    #topborders
    ws['B1'].border = top_thin_border
    ws['C1'].border = top_thin_border
    ws['D1'].border = top_thin_border
    ws['E1'].border = top_thin_border
    ws['D2'].border = top_thin_border
    ws['E2'].border = top_thin_border
    ws['D3'].border = top_thin_border
    ws['E3'].border = top_thin_border
    #top_right borders
    ws['F1'].border = top_right_thin_border
    ws['F2'].border = top_right_thin_border
    ws['F3'].border = top_right_thin_border

    '''
    info blocks
    '''
    ws['B2'] = "Job #: " + h2
    ws['B3'] = B3
    ws['A5'] = A5
    ws['H4'] = "Job #: " + h2
    ws['H5'] = "Program: " + H5
    ws['H6'] = "Supplier: " + H6
    ws['H7'] = "Buyer: " + H7 + ' (D' + H71 + ')'
    ws['H8'] = "Artwork due date: " + H8
    ws['H9'] = "Packout date: " + H9
    ws['H10'] = "Shipdate: " + H10
    ws['H11'] = "In-store date: " + H11
    ws['H13'] = "联系人: " + H13

    wb.save(path1)
    return


if __name__ == "__main__":
    run()




