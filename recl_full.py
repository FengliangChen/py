#!/usr/bin/env python3

# encoding: utf-8
# BY Simon Chen, bafelem@gmail.com
# Date: 07/10/2018
# Tool helping to fill the job sheet.

import datetime
import re
from openpyxl import load_workbook
import sys

path1 = sys.argv[2]

h2 = sys.argv[1]
h2 = h2.upper()

b31 = input("Please enter the season code:")
b31 = b31.upper()

b32 = input("Please enter country, A. China, B. USA, C.Canada, otherwise China:")
if re.match('A', b32, re.IGNORECASE ):
	b32 = 'China'
elif re.match('B', b32, re.IGNORECASE ):
	b32 = 'USA'
elif re.match('C', b32, re.IGNORECASE ):
	b32 = 'Canada'
else:
	b32 = 'China'
	print("It will default to China.")

B3 = b31 + '/' + b32


A5 = input("Please enter the date or default today:")
if A5.strip() == "":
	time = datetime.datetime.now()
	time = int(time.hour)
	today = datetime.date.today()
	if time >= 7:
		A5 = today.strftime( '%m/%d/%Y')
	else:
		A5 = today + datetime.timedelta( days = -1 )
		A5 = A5.strftime( '%m/%d/%Y')

# Dictionary of full seasional names.
dic = {'THX':'Thanksgiving', 'AW':'Athletic Works', 'NY':'New Year', 'HOL':'Holiday Time', 'EA':'Easter', 'SUM':'Summer', 'SP':'Spring', 'CASE':'Casemate', 'VAL':'Valentines', 'ONN':'ONN', 'HV':'Harvest', 'HW':'Halloween', 'PD':'Play Day', 'GD':'Graduation', 'MO':'Mother\'s Day', 'AMR':'Americana', 'GV':'Great Value', 'STP':'St. Patrick', 'MG':'Mardi Gras', 'BW':'Blackweb', 'HAN':'Hanukkah', 'WTC':'Way to Celebrate', 'FA':'Father\'s Day', 'PC':'Parent\'s Choice', 'CA':'Canadiana', 'EQ':'Equate', 'MOV':'Movelo', 'WTCW':'Way To Celebrate Wedding'}

try:
	code = b31
	full_name = dic[code]
except KeyError:
	full_name = input("Please enter full season name:")
finally:
	H5 = full_name

#other info, will change to title case.
H6 = input("Please enter supplier name:")
H6 = H6.title()

H7 = input("Please enter buyer name:")
H7 = H7.title()
H71 = input("Please enter Department number:")

H11 = input("In-store date:")
H10 = input("Shipdate:")
H9 = input("Packout date:")
H8 = input("Please Artwork due date:")
H13 = input("Contact info:")


# openlxs blocks
wb = load_workbook(path1)
ws = wb.active

ws['B2'] = "Job #: " + h2
ws['B3'] = B3
ws['A5'] = A5
ws['H4'] = "Job #: " + h2
ws['H5'] = "Program: " + H5
ws['H6'] = "Supplier: " + H6
ws['H7'] = "Buyer: " + H7 + ' (D' + H71 + ')'
ws['H8'] = "Artwork due date: " + H8
ws['H9'] = "Packout date: " + H9
ws['H10'] = "Shipdate: " + H10
ws['H11'] = "In-store date: " + H11
ws['H13'] = "联系人: " + H13

wb.save(path1)

print("Completed! Enjoy!")

exit()